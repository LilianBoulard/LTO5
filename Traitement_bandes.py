"""

Ce script a pour rôle de traiter les fichiers enregistrés par le script "Sauvegarde des bandes" en les analysant, sortant les informations utiles, et sauvegardant celles-ci dans un fichier de suivi.

"""

import sys

if sys.version_info[0] != 3: # Si la version de Python n'est pas la 3.x
    input('Ce script doit être exécuté avec Python 3')
    exit(0)

import calendar, datetime



# ==========================
#  Définition des variables
# ==========================


# Nom de la boite mail dans lequel le script de sauvegarde va aller chercher les mails avec les objets ci-dessous.
nomBoite            = "adresse.email@domaine.exemple"
# Objets des mails à sauvegarder, sous la forme d'une liste.
objets              = ["Bandes_LTO5_FREEPOOL", "Bandes_LTO5_FREEPOOL_COFFRE"]

# Le fichier dans lequel seront insérées les données. Peut être un chemin relatif ou absolut.
fichierPrincipal    = 'Bandes_LTO5_FREEPOOL.xlsx'
# Dossier dans lequel se trouvent les dossiers dans lesquels se trouvent les fichiers ci-dessous.
# Notez bien le / à la fin, il est important. Peut être un chemin relatif ou absolut.
dossierBandes       = '//Serveur/LTO/Suivi/Fichiers/'
# Extension des fichiers ci-dessous. Si cette extension doit changer de type (xls, csv, etc), il faut modifier la méthode SaveAs plus bas. Voir la note dans le script de sauvegarde.
extension           = '.htm'
# Format du fichier de bandes pour la baie. %s correspond à la date du jour au format "formatDate" (ci-dessous), cette expression doit rester telle quelle même si elle peut changer d'emplacement.
fichierBaie         = 'Bandes_LTO5_FREEPOOL-%s' + extension
# Format du fichier de bandes pour le coffre. Pareil ici, laisser le %s là où est censée se trouver la date.
fichierCoffre       = 'Bandes_LTO5_FREEPOOL_COFFRE-%s' + extension
# Utilisé dans le nommage des fichiers ci-dessus (%s) ; Voir https://docs.python.org/2/library/datetime.html#strftime-and-strptime-behavior
formatDate          = '%d%m%Y'
# Intitulés des colonnes du fichier principal
colonnes            = ['LTO5_FREEPOOL', 'Nombre de bandes ajoutées', 'LTO5_GPS_FREEPOOL', 'Nombre de bandes ajoutées GPS', 'Coffre', 'Nombre de bandes retirées', 'Total']
# Liste des choix "positifs". Insensible à la casse.
choixPositifs       = ['y', 'ye', 'yes', 'oui', 'ou', 'o', 'ok']
# Active ou désactive le mode debug
verbose             = False

# On évitera les accents et autres caractères spéciaux
jours = [
    'Lundi',
    'Mardi',
    'Mercredi',
    'Jeudi',
    'Vendredi',
    'Samedi',
    'Dimanche'
]
mois = [
    'Janvier',
    'Fevrier',
    'Mars',
    'Avril',
    'Mai',
    'Juin',
    'Juillet',
    'Aout',
    'Septembre',
    'Octobre',
    'Novembre',
    'Decembre'
]



# ================================
#  Import des modules nécessaires
# ================================



def install_module(package):
    """
    Utilise pip pour installer un nouveau module Python.
    Retourne un booléen "Vrai" lorsque ça fonctionne, "Faux" autrement.
    """
    import subprocess, sys
    try:
        # Lance la commande pip pour installer le module
        if subprocess.check_call([sys.executable, "-m", "pip", "install", package]) == 0:
            return True
    except subprocess.CalledProcessError as e:
        input("Erreur rencontrée lors de l'installation du module \"{}\" : {}\nMerci de l'installer manuellement en utilisant pip.".format(package, e))
        return False



try:
    import openpyxl
except ImportError:
    if install_module("openpyxl"):
        import openpyxl
    else:
        exit()

try:
    import pandas
except ImportError:
    if install_module("pandas"):
        import pandas
    else:
        exit()

try:
    from bs4 import BeautifulSoup
except ImportError:
    if install_module("beautifulsoup4"):
        from bs4 import BeautifulSoup
    else:
        exit()

try:
    import lxml
except ImportError:
    if install_module("lxml"):
        import lxml
    else:
        exit()

try:
    import xlrd
except ImportError:
    if install_module("xlrd"):
        import xlrd
    else:
        exit()

try:
    import html5lib
except ImportError:
    if install_module("html5lib"):
        import html5lib
    else:
        exit()

"""
try:
    import jinja2
except ImportError:
    if install_module("Jinja2"):
        import jinja2
    else:
        exit()
"""



# ======
#  Code
# ======



def main(dossierBandes):
    """
    Fonction principale
    """

    print("TRAITEMENT DES BANDES\n\n\
Ce programme permet d'insérer automatiquement les valeurs des fichiers de bandes dans un fichier principal.\n\
Veuillez consulter le code source pour plus d'informations.\n\n")

    choixTemps = input("Appuyez sur Entrée pour utiliser la date d'aujourd'hui.\nSinon, entrez une date au format " + '"AAAA, MM, JJ"\nIl est aussi possible de donner une plage en séparant deux dates avec ; . Si trois dates ou plus sont entrées, le script traitera uniquement ces dates.\n> ')

    # Initialise la liste "tempsListe"
    tempsListe = []

    # Si le choix de la date n'est pas vide
    if choixTemps != '':
        # Découpe l'input à partir du caractère ";" : sépare les dates entrées.
        choixTemps = choixTemps.split(';')

        # Pour chaque date dans la liste entrée par l'utilisateur
        for choix in choixTemps:
            # On récupère l'index de la boucle for
            index = choixTemps.index(choix)
            # Divise chaque date de la liste depuis "," afin de contenir en index 0 l'année, index 1 le mois et index 2 le jour.
            choixTemps[index] = choixTemps[index].split(',')
            # choixTemps est maintenant une liste multi-dimensionnelle (listes dans une liste).

        # S'il y a exactement deux dates dans la liste, alors on veux les utiliser comme une plage
        if len(choixTemps) == 2:
            # Si les mois correspondent, et que la première date est plus petite (récente) que la deuxième:
            if choixTemps[0][1] == choixTemps[1][1] and choixTemps[1][2] > choixTemps[0][2]:
                for i in range(int(choixTemps[0][2]), (int(choixTemps[1][2]) + 1), 1): # Pour chaque jour entre la première et la deuxième date
                    try:
                        # On essaye d'ajouter la date (sous forme d'objet datetime, plus facile à manipuler) à la liste "tempsListe"
                        tempsListe.append(datetime.datetime(int(choixTemps[0][0]), int(choixTemps[0][1]), i, 00, 00))
                    except (ValueError, IndexError): # Si ça ne fonctionne pas, c'est que la date n'est pas valide
                        input('Erreur: Date invalide, merci de la vérifier et de réessayer.')
                        exit(0)
            else:
                print("Invalide. Le mois doit être le même et les dates doivent être dans l'ordre.")
                input('Appuyez sur Entrée pour fermer le programme.')
                exit(0)
        elif len(choixTemps) == 1: # Si il n'y a qu'une seule date
            try:
                # On ajoute la date sous forme d'objet datetime à la liste
                tempsListe.append(datetime.datetime(int(choixTemps[0][0]), int(choixTemps[0][1]), int(choixTemps[0][2]), 00, 00))
            except (ValueError, IndexError):
                input('Erreur: Date invalide, merci de la vérifier et de réessayer.')
                exit(0)
        else: # S'il y a plus de deux dates
            for temps in choixTemps: # Pour chaque date dans la liste
                # On récupère l'index de la boucle for
                index = choixTemps.index(temps)
                try:
                    # On ajoute la date sous forme d'objet datetime à la liste
                    tempsListe.append(datetime.datetime(int(choixTemps[index][0]), int(choixTemps[index][1]), int(choixTemps[index][2]), 00, 00))
                except (ValueError, IndexError):
                    input('Erreur: Date invalide, merci de la vérifier et de réessayer.')
                    exit(0)
    else: # Si le choix de la date est vide
        # On utilise la date d'aujourd'hui
        tempsListe.append(datetime.datetime.now())

        print('Date sélectionnée : ')
        for temps in tempsListe: # Pour chaque date
            # Appelle la fonction getDay(), qui retourne une date au formoat "Jeudi 1 janvier 1970"
            print(getDay(tempsListe[0], tempsListe[0].day))

    # On initialise le choix 1 pour permettre par la suite l'utilisation d'un onglet par défaut, et éviter au script de redemander quel onglet utiliser à chaque itération (quand on utilise une plage de dates par exemple)
    ch1 = 'NULL'

    for temps in tempsListe: # Pour chaque date dans la liste
        # Construit le chemin où le script ira chercher les fichiers de bandes
        repertoire = (dossierBandes + mois[temps.month - 1] + str(temps.year)) + '/'

        # Formatte la date selon "formatDate" (un format de date donc). Cette variable est initialisée en en-tête du script
        formattedDate = temps.strftime(formatDate)
        # On utilise la date formattée au dessus pour former le nom du fichier baie
        formattedFichierBaie = repertoire + fichierBaie % (formattedDate)
        # Et du fichier coffre
        formattedFichierCoffre = repertoire + fichierCoffre % (formattedDate)

        # Construit le nom de l'onglet principal (du fichier principal) en fonction de la date (peut être modifié par l'utilisateur par la suite)
        ongletFichierPrincipal = mois[(temps.month - 1)] + str(temps.year)

        if ch1 == 'NULL': # Si ch1 est initialisé (égal à NULL)
            # On demande quel onglet utiliser
            ch1 = input("Utiliser l'onglet par défaut ? (%s) Si non, entrez le nom à utiliser\n> " % (ongletFichierPrincipal))
        if ch1 != '': # Si l'utilisateur ne laisse pas le champ vide
            # On défini son choix (la phrase qu'il a entré) comme nouvel onglet principal
            ongletFichierPrincipal = ch1

        if ongletFichierPrincipal in pandas.ExcelFile(fichierPrincipal).sheet_names: # Si l'onglet principal existe dans le fichier principal
            try:
                # On essaye d'écrire une ligne (appelle la fonction ecrire_ligne())
                ecrire_ligne(fichierPrincipal, ongletFichierPrincipal, formattedFichierBaie, formattedFichierCoffre, temps)
            except PermissionError: # On attrape les erreurs de permission
                print('Permission refusée. Merci de fermer le fichier et de réessayer.')
                break
            except FileNotFoundError: # Et les erreurs de type fichier non trouvé
                print("Un fichier n'a pas été trouvé.")
                break
        else: # Si l'onglet n'existe pas dans le fichier
            # On demande si on doit le créer
            ch2 = input("L'onglet %s n'existe pas. Le créer ? [Oui/Non] " % (ongletFichierPrincipal))
            if ch2.lower() in choixPositifs: #Si la réponse de l'utilisateur se trouve dans la liste des réponses positives
                try:
                    # On essaye de construire le nouvel onglet
                    creer_nouvel_onglet(fichierPrincipal, ongletFichierPrincipal, temps)
                except PermissionError:
                    print('Permission refusée ; merci de fermer le fichier et de réessayer.')
                    break
                except FileNotFoundError:
                    print("Un fichier n'a pas été trouvé. Fermeture du programme.")
                    break
            else: # Si la réponse de l'utilisateur ne se trouve pas dans la liste
                print('Réponse négative donnée.')
                break
    input('Appuyez sur Entrée pour fermer le programme.')


def ecrire_ligne(fichierPrincipal, ongletFichierPrincipal, fichierBaie, fichierCoffre, temps):
    """
    Cette fonction écrit des données venant de "fichierBaie" et "fichierCoffre" dans l'onglet "ongletFichierPrincipal" du fichier "fichierPrincipal"
    """

    # Appelle la fonction "compter_nombre_bandes()", qui récupère le nombre de bandes dans la baie.
    bandesFichierBaie = compter_nombre_bandes(fichierBaie)
    # Pareil ici pour le nombre de bandes dans le coffre.
    bandesFichierCoffre = compter_nombre_bandes(fichierCoffre)

    LTO5_FREEPOOL = bandesFichierBaie[0]
    LTO5_GPS_FREEPOOL = bandesFichierBaie[1]
    Coffre = bandesFichierCoffre[0] + bandesFichierCoffre[1]
    Total = bandesFichierBaie[0] + bandesFichierBaie[1] + bandesFichierCoffre[0] + bandesFichierCoffre[1]

    # Si la fonction "getTapesCount" appelée ci-dessusa retourné 1337, qui est un code d'erreur (ce n'est pas une norme, il peut être changé au besoin),
    if bandesFichierBaie[0] == 1337 or bandesFichierCoffre[0] == 1337:
        # On arrête la fonction et on ne retourne rien.
        return

    # On charge le fichier
    book = openpyxl.load_workbook(fichierPrincipal)
    # On initialise le writer en lui donnant le nom du fichier à ouvrir.
    writer = pandas.ExcelWriter(fichierPrincipal, engine='openpyxl')
    # Et on fournit l'objet openpyxl (le fichier) au writer
    writer.book = book

    # Stocke le contenu de l'onglet "ongletFichierPrincipal" du fichier "fichierPrincipal", dans la variable "df", sous la forme d'un DataFrame pandas.
    df = pandas.read_excel(fichierPrincipal, ongletFichierPrincipal)
    df.loc[(temps.day - 1)] = pandas.Series({
        colonnes[0]:LTO5_FREEPOOL, # Colle le contenu de la variable LTO5_FREEPOOL dans la colonne 0 (0 car la colonnes index - les dates - n'est pas compris dans le DataFrame).
        colonnes[1]:df.iat[(temps.day - 1), 2], # Récupère le contenu de la cellule et le recolle au même endroit.
        colonnes[2]:LTO5_GPS_FREEPOOL,
        colonnes[3]:df.iat[(temps.day - 1), 4],
        colonnes[4]:Coffre,
        colonnes[5]:df.iat[(temps.day - 1), 6],
        colonnes[6]:Total
    }) # Construit le DataFrame

    if verbose == True: # Si le mode debug est activé (voir en-tête).
        # On print la première ligne du DataFrame.
        print(df.index)

    # On supprime la première colonne. Une colonne en trop est créée à chaque fois qu'on load un excel.
    df = df.drop(df.columns[0], 1)
    # On défini la première colonne du DataFrame.
    df.columns = [colonnes]

    # On récupère les lignes.
    lignes = getIndexes(temps)
    if len(df.index) != len(lignes): # Si le nombre de lignes du DataFrame et de la liste ne sont pas égales
        # On calcule la dif (qui est forcément un chiffre supérieur à 0, d'où la fonction abs())
        dif = abs(len(df.index) - (len(lignes)))
        while dif != 0: # Tant que la dif n'est pas à 0
            # On supprime la première ligne.
            df = df.drop(df.index[0])
            # Et on décrémente la différence.
            dif -= 1

    # Construction de la première ligne.
    df.index = [lignes]

    # Passe les onglets existants au writer. Sans cette ligne, les onglets non appelés explicitement seraient supprimés.
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # Envoie les données au fichier Excel.
    df.to_excel(writer, ongletFichierPrincipal)
    # Et le sauvegarde.
    writer.save()

    # Réimporte le fichier principal.
    wb = openpyxl.load_workbook(fichierPrincipal)
    # Et supprime sa première ligne (un bug qu'il n'est pas possible de réparer depuis la première instance...).
    wb[ongletFichierPrincipal].delete_rows(2, 1)
    # Et le sauvegarde.
    wb.save(fichierPrincipal)

    # Résume les informations.
    print('Données ajoutées avec succès au fichier pour le {0}.\nBandes | LTO5: {1}, GPS: {2}, Coffre: {3}, Total: {4}\n'.format(getDay(temps, (temps.day)), LTO5_FREEPOOL, LTO5_GPS_FREEPOOL, Coffre, Total))


def compter_nombre_bandes(fichier):
    """
    Retourne le nombre de bandes trouvées dans le fichier "fichier"
    """

    try:
        # Récupère les contenu des balises <table> du fichier html/htm.
        bandes = pandas.read_html(fichier)[1]
        # Récupération du contenu de la colonne 7, et incrémentation de la variable si le contenu est égal à "LTO5_FREEPOOL"
        bandes_LTO5_FREEPOOL = [x for x in bandes[7] if x == "LTO5_FREEPOOL"]
        # Ici aussi, mais pour "LTO5_GPS_FREEPOOL"
        bandes_LTO5_GPS_FREEPOOL = [x for x in bandes[7] if x == "LTO5_GPS_FREEPOOL"]
        # On retourne dans un tuple le nombre de bandes LTO5_FREEPOOL et LTO5_GPS_FREEPOOL.
        return (len(bandes_LTO5_FREEPOOL), len(bandes_LTO5_GPS_FREEPOOL))
    except (FileNotFoundError, ValueError):
        print("Impossible d'importer le fichier %s - merci de le télécharger." % (fichier))
        # Retourne un code d'erreur, que j'ai défini (ce n'est pas une norme, il peut être changé au besoin).
        return (1337, 1337)


def creer_nouvel_onglet(fichierPrincipal, ongletFichierPrincipal, temps):
    """
    Créer un nouvel onglet "ongletFichierPrincipal" dans le fichier "fichierPrincipal" et initialise la première ligne et la première colonne
    """

    # On ouvre le fichier Excel.
    book = openpyxl.load_workbook(fichierPrincipal)
    # On initialise le writer.
    writer = pandas.ExcelWriter(fichierPrincipal, engine='openpyxl')
    # Et on lui fournit le fichier excel.
    writer.book = book

    df=pandas.DataFrame(
        columns=[colonnes],
        index=[getIndexes(temps)]
    ) # Construit le DataFrame, en y incluant la première ligne et la première colonne.
    # Modifie le style du DataFrame, nécessire le module Janja2.
    #df.style.set_properties({'text-align': 'right'})

    # Passe les onglets existants au writer. Sans cette ligne, les onglets non appelés explicitement seraient supprimés.
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # On écrit les données dans le fichier Excel.
    df.to_excel(writer, ongletFichierPrincipal)
    # Et on le sauvegarde.
    writer.save()

    print('Nouvel onglet "%s" créé avec succès.' % (ongletFichierPrincipal))


def getIndexes(temps):
    """
    Retourne une liste de tous les jours du mois au format "Jeudi 01 Janvier 1970"
    """

    # Récupère une liste des jours du mois dont fait parti la date "temps"
    monthDays = calendar.Calendar().itermonthdays(temps.year, temps.month)
    # Initialisation de la liste
    days = []
    for monthDay in monthDays: # Pour chaque jour dans la liste des jours du mois
        if monthDay != 0 & monthDay != '': # Si le jour n'est pas vide ou égal à zéro (il est courant que la fonction retourne des champs vides)
            if verbose == True: # Si le mode debug est activé (voir en-tête)
                print('[getIndexes] Monthday is :' + str(monthDay))
            # Ajoute à la liste une date, retournée par la fonction getDay()
            days.append(getDay(temps, monthDay))
    # Retourne la liste des jours, sous le format "Jeudi 01 Janvier 1970"
    return days


def getDay(temps, jour):
    """
    Retourne une date au format "Jeudi 01 janvier 1970".
    """

    return '{jour} {nbJour} {mois} {an}'.format(
        jour = jours[datetime.datetime.weekday(datetime.datetime.strptime("{0}/{1}/{2}".format(
            jour,
            temps.month,
            temps.year
        ), "%d/%m/%Y"))],
        nbJour = jour,
        mois = mois[temps.month - 1],
        an = temps.year
    )



# Si le fichier est ouvert directement
if __name__ == '__main__':
    main(dossierBandes)
